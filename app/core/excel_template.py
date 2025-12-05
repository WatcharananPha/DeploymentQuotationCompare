# app/core/excel_template.py
from __future__ import annotations

import os
import tempfile
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from .processing import (
    COMPANY_NAME_ROW,
    CONTACT_INFO_ROW,
    HEADER_ROW,
    ITEM_MASTER_LIST_COL,
    COLUMNS_PER_SUPPLIER,
    SUMMARY_LABELS,
    update_google_sheet_for_single_file,
)

class ExcelWorksheetAdapter:
    def __init__(self, ws):
        self.ws = ws

    @property
    def col_count(self) -> int:
        return self.ws.max_column

    def get_all_values(self) -> List[List[str]]:
        max_row = self.ws.max_row
        max_col = self.ws.max_column
        values: List[List[str]] = []
        for r in range(1, max_row + 1):
            row_vals: List[str] = []
            for c in range(1, max_col + 1):
                v = self.ws.cell(row=r, column=c).value
                row_vals.append("" if v is None else str(v))
            values.append(row_vals)
        return values

    def insert_rows(self, rows: List[List[Any]], row_index: int) -> None:
        count = len(rows)
        for i in range(count):
            self.ws.insert_rows(row_index + i)

    def batch_update(self, batch_requests: List[Dict[str, Any]], value_input_option: str = "USER_ENTERED") -> None:
        for req in batch_requests:
            rng = req["range"]
            vals = req["values"]
            if ":" in rng:
                start_ref, end_ref = rng.split(":")
            else:
                start_ref = rng
                end_ref = rng

            start_col_letters = "".join(ch for ch in start_ref if ch.isalpha())
            start_row_digits = "".join(ch for ch in start_ref if ch.isdigit())
            end_col_letters = "".join(ch for ch in end_ref if ch.isalpha())
            end_row_digits = "".join(ch for ch in end_ref if ch.isdigit())

            start_col = column_index_from_string(start_col_letters)
            start_row = int(start_row_digits)
            end_col = column_index_from_string(end_col_letters)
            end_row = int(end_row_digits)

            for r_offset, row_vals in enumerate(vals):
                row_idx = start_row + r_offset
                if row_idx > end_row:
                    break
                for c_offset, value in enumerate(row_vals):
                    col_idx = start_col + c_offset
                    if col_idx > end_col:
                        break
                    self.ws.cell(row=row_idx, column=col_idx).value = value

def _resolve_template_path() -> Path:
    base_dir = Path(__file__).resolve().parent.parent
    templates_dir = base_dir / "templates"
    root_dir = base_dir.parent
    candidates = [
        templates_dir / "quotation_template.xlsx",
        templates_dir / "temp.xlsx",
        root_dir / "quotation_template.xlsx",
        root_dir / "temp.xlsx",
    ]
    for path in candidates:
        if path.exists():
            return path
    return root_dir / "temp.xlsx"

def generate_excel_from_results(results: List[Dict[str, Any]]) -> str:
    template_path = _resolve_template_path()
    wb = load_workbook(template_path)
    ws = wb.active
    adapter = ExcelWorksheetAdapter(ws)

    initial_sheet_values = adapter.get_all_values()

    live_existing_products: List[Dict[str, Any]] = []
    for row_idx, row in enumerate(initial_sheet_values[HEADER_ROW:], start=HEADER_ROW + 1):
        if len(row) >= ITEM_MASTER_LIST_COL:
            name = row[ITEM_MASTER_LIST_COL - 1].strip()
            if name and name not in SUMMARY_LABELS:
                live_existing_products.append({"name": name, "row": row_idx})

    live_existing_suppliers: Dict[str, int] = {}
    header_row_values = initial_sheet_values[COMPANY_NAME_ROW - 1] if initial_sheet_values else []
    for col_idx in range(ITEM_MASTER_LIST_COL + 1, len(header_row_values) + 1, COLUMNS_PER_SUPPLIER):
        supplier_name = header_row_values[col_idx - 1].strip() if (col_idx - 1) < len(header_row_values) else ""
        if supplier_name:
            live_existing_suppliers[supplier_name] = col_idx

    for data in results:
        if data:
            live_existing_products, live_existing_suppliers = update_google_sheet_for_single_file(
                adapter, data, live_existing_products, live_existing_suppliers
            )

    fd, output_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(output_path)
    return output_path